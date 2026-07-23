#!/usr/bin/env python3
"""Build an analyst-friendly quiz-data.xlsx from responses.csv + feedback.csv.

Tabs: Responses · Feedback · Summary · Question Key
Run:  python3 build_sheet.py
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
GREEN = "163F35"
GREEN_SOFT = "E7EEEB"

def read_csv(name):
    path = os.path.join(BASE, name)
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def fmt_ts(iso):
    # 2026-07-22T21:10:22.210Z -> 2026-07-22 21:10 UTC
    if not iso:
        return ""
    try:
        d, t = iso.replace("Z", "").split("T")
        return f"{d} {t[:5]} UTC"
    except Exception:
        return iso

COND_LABEL = {
    "cond-energy": "Energy + focus",
    "cond-gut": "Gut health + comfort",
    "cond-stress": "Stress, sleep + mood",
    "cond-beauty": "Beauty, aging + longevity",
    "cond-other": "Overall wellness + immunity",
}

# (source key, friendly header, width, wrap?)
RESP_COLS = [
    ("name", "Name", 16, False),
    ("timestamp", "Submitted", 20, False),
    ("focus", "Primary focus (Q1)", 22, True),
    ("wishlist", "Wellness wishlist", 30, True),
    ("feeling", "Body feeling lately", 30, True),
    ("barriers", "Barriers to feeling best", 34, True),
    ("experience", "Supplement experience", 30, True),
    ("routine_now", "Current routine", 26, True),
    ("flags", "Health flags", 30, True),
    ("commitment", "Routine commitment", 26, True),
    ("begin", "Prefers to begin", 28, True),
    ("mindset", "Wellness mindset", 32, True),
    ("dd_energy", "Deep-dive · Energy + focus", 26, True),
    ("dd_gut", "Deep-dive · Gut health", 26, True),
    ("dd_stress", "Deep-dive · Stress/sleep/mood", 26, True),
    ("dd_beauty", "Deep-dive · Beauty/aging", 26, True),
    ("dd_other", "Deep-dive · Overall/other", 26, True),
    ("submission_id", "Submission ID", 18, False),
]

FB_COLS = [
    ("name", "Name", 16, False),
    ("timestamp", "Submitted", 20, False),
    ("rating", "Rating (1-5)", 12, False),
    ("ease", "Ease of use", 18, True),
    ("comment", "Comment", 50, True),
    ("submission_id", "Submission ID", 18, False),
]

QUESTION_KEY = [
    ("focus", "What would you love a little more support with right now?", "Single", "Energy + focus | Gut health + comfort | Stress, sleep + mood | Beauty, aging + longevity | Overall wellness + immunity"),
    ("wishlist", "What else is on your wellness wishlist?", "Multi", "Same five focus areas"),
    ("feeling", "How has your body been feeling lately?", "Single", "Energized and balanced | Getting through the day… | Stressed, tired, or overwhelmed | Managing recurring challenges | Not totally sure"),
    ("barriers", "What tends to get in the way of feeling your best?", "Multi", "Energy | Stress/sleep | Digestive discomfort | Not sure where to start | Busy life | Budget"),
    ("experience", "How would you describe your supplement experience?", "Single", "Just getting started | A few basics | Consistent routine | Tried several | Practitioner recommended"),
    ("routine_now", "What does your current routine look like?", "Single", "None | 1–2 products | 3–5 products | 6+ products | Inconsistent"),
    ("flags", "Anything important we should know before we build your plan?", "Multi", "Sensitive stomach | Pregnant/breastfeeding | On medications | Wants practitioner | None"),
    ("commitment", "What kind of routine would you actually stick with?", "Single", "One simple product | A few essentials | A complete routine | Start small & build"),
    ("begin", "How would you prefer to begin?", "Single", "One-time first | Subscribe if benefits clear | Subscribe for consistency | Not sure"),
    ("mindset", "What best describes your wellness mindset right now?", "Single", "Feel better daily | Long-term routine | Prevention | Recovering from stress | Optimizing | Understanding my body"),
    ("conditional_topic", "Deep-dive (branches off Q1 primary focus)", "Multi", "Shown for whichever focus is picked in Q1; answers are the topic-specific symptoms/timing"),
]

# ---- styles ----
hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=GREEN)
cell_font = Font(name="Arial", size=10)
top_align = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    ws.row_dimensions[1].height = 30

def write_table(ws, cols, rows, transform=None):
    for i, (_k, header, width, _wrap) in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=header)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, len(cols))
    for r, row in enumerate(rows, start=2):
        data = transform(row) if transform else row
        for i, (key, _h, _w, wrap) in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i, value=data.get(key, ""))
            cell.font = cell_font
            cell.alignment = top_align if wrap else Alignment(vertical="top")
            cell.border = border
        # zebra striping
        if r % 2 == 0:
            for i in range(1, len(cols) + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GREEN_SOFT)

def resp_transform(row):
    d = dict(row)
    d["timestamp"] = fmt_ts(row.get("timestamp", ""))
    return d

def fb_transform(row):
    d = dict(row)
    d["timestamp"] = fmt_ts(row.get("timestamp", ""))
    return d

def main():
    responses = read_csv("responses.csv")
    feedback = read_csv("feedback.csv")

    wb = Workbook()

    # Responses
    ws = wb.active
    ws.title = "Responses"
    write_table(ws, RESP_COLS, responses, resp_transform)

    # Feedback
    wf = wb.create_sheet("Feedback")
    write_table(wf, FB_COLS, feedback, fb_transform)

    # Summary
    wsum = wb.create_sheet("Summary")
    wsum.column_dimensions["A"].width = 30
    wsum.column_dimensions["B"].width = 16
    wsum["A1"] = "Cymbiotika Quiz — Summary"
    wsum["A1"].font = Font(name="Arial", size=13, bold=True, color=GREEN)
    rows = [("Metric", "Value"),
            ("Total completed quizzes", len(responses)),
            ("Total feedback submissions", len(feedback))]
    ratings = [int(f["rating"]) for f in feedback if f.get("rating", "").isdigit()]
    if ratings:
        rows.append(("Average rating", round(sum(ratings) / len(ratings), 2)))
    rows.append(("", ""))
    rows.append(("Primary focus (Q1)", "Count"))
    focus_counts = {}
    for r in responses:
        focus_counts[r.get("focus", "")] = focus_counts.get(r.get("focus", ""), 0) + 1
    for k in sorted(focus_counts, key=lambda x: -focus_counts[x]):
        if k:
            rows.append((k, focus_counts[k]))
    for r, (a, b) in enumerate(rows, start=3):
        ca, cb = wsum.cell(row=r, column=1, value=a), wsum.cell(row=r, column=2, value=b)
        if b == "Value" or b == "Count":
            ca.font = hdr_font; ca.fill = hdr_fill
            cb.font = hdr_font; cb.fill = hdr_fill
        else:
            ca.font = cell_font; cb.font = cell_font

    # Question Key
    wk = wb.create_sheet("Question Key")
    kcols = [("id", "Field / column", 22, False), ("q", "Question", 55, True),
             ("type", "Type", 10, False), ("opts", "Options", 70, True)]
    for i, (_k, h, w, _wrap) in enumerate(kcols, start=1):
        wk.cell(row=1, column=i, value=h)
        wk.column_dimensions[get_column_letter(i)].width = w
    style_header(wk, len(kcols))
    for r, (fid, q, typ, opts) in enumerate(QUESTION_KEY, start=2):
        for i, val in enumerate([fid, q, typ, opts], start=1):
            cell = wk.cell(row=r, column=i, value=val)
            cell.font = cell_font
            cell.alignment = top_align
            cell.border = border

    out = os.path.join(BASE, "quiz-data.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(responses)} responses, {len(feedback)} feedback)")

if __name__ == "__main__":
    main()
